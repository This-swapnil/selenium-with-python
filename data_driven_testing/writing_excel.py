from os import path
import openpyxl

path = "Data/writing_data_3.xlsx"
workbook = openpyxl.load_workbook(path)

sheet = workbook.active

# Same data
"""
for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "welcome
"""

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Smith"

sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "John"

sheet.cell(3, 1).value = 789
sheet.cell(3, 2).value = "David"


workbook.save(path)
