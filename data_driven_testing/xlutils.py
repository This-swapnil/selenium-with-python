import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(rownum, colnum).value


def write_data(file, sheet_name, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(rownum, colnum).value = data
    workbook.save(file)


def fill_green_color(file, sheet_name, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(
        start_color="60b212", end_color="60b212", fill_type="solid"
    )
    sheet.cell(rownum, colnum).fill = green_fill
    workbook.save(file)


def fill_red_color(file, sheet_name, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    sheet.cell(rownum, colnum).fill = red_fill
    workbook.save(file)
