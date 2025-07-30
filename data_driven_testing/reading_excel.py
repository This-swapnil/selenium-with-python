import openpyxl

# Path => Workbook => Sheets => Rows => Cells

path = "Data/data.xlsx"
workbook = openpyxl.load_workbook(path)

sheet = workbook["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column

for i in range(1, rows + 1):
    for j in range(1, cols + 1):
        print(sheet.cell(i, j).value, end=" ")
    print()
